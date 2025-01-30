from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
import win32com.client as win32
import time
import datetime
from selenium.common.exceptions import WebDriverException
import os



def ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ):
    # Credentials to log in to the website
    username = 'sabrine.hassairi.adm'
    password = 'Ningen0925*'
    # URL of the website to scrape
    url = 'https://10.6.4.120/dashboards/custom/22'
    url1 = 'https://10.6.4.120/dashboards/custom/23'
    # Set up the Chrome driver service
    service = webdriver.chrome.service.Service('/Applications/driver/chromedriver')
    # Set up the webdriver
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--ignore-ssl-errors')
    driver = webdriver.Chrome(service=service, options=options)

    # Navigate to the login page
    driver.get('https://10.6.4.120/login')
    time.sleep(10)
    # insert the login details
    username_field = driver.find_element(By.XPATH,
                                         "/html/body/div[1]/div/md-content/div/div[1]/div/form/md-input-container[1]/input")
    username_field.send_keys(username)
    password_field = driver.find_element(By.XPATH,
                                         "/html/body/div[1]/div/md-content/div/div[1]/div/form/md-input-container[2]/input")
    password_field.send_keys(password)
    # Submit the login form
    password_field.send_keys(Keys.RETURN)
    # Wait for the login to complete
    time.sleep(10)



    # Navigate to the URL1
    driver.get(url1)
    # Wait for the page to load
    time.sleep(10)
    xpaths1 = [

        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[1]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[2]/md-card/md-card-content/div/div/div/div/div[1]"
    ]
    value3, value4 = [element.text for element in
                      driver.find_elements(By.XPATH, " | ".join(xpaths1))]

    CS_ODS_DZ = value3
    CS_ODS_TR_DZ = value4

    print(CS_ODS_DZ)
    print(CS_ODS_TR_DZ)

    # Load the Excel file
    ######## modifier : nom du fichier excel par la version récente #########################################################################"
    we = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")

    # Select the worksheet you want to work with
    ws = we.active

    # Select the cell to write in
    cell = ws.cell(row=row_ODS_DZ, column=column_ODS_DZ)
    lastval5 = int(CS_ODS_DZ) - somme_ODS_DZ_recu
    # write the new value
    cell.value = lastval5
    # new value of sum
    somme_ODS_DZ_recu += lastval5

    # Select the cell to write in
    cell = ws.cell(row=row_ODS_DZ, column=column_ODS_DZ_tr)
    lastval5 = int(CS_ODS_TR_DZ) - somme_ODS_DZ_tr
    # write the new value
    cell.value = lastval5
    # new value of sum
    somme_ODS_DZ_tr += lastval5

    ######## modifier : nom du fichier excel par la version récente################################################################################
    we.save(r"C:\Users\Melek\Desktop\QS1.xlsx")

    # set the number of times to send the email
    num_emails = 1

    # create Dispatch objects for Excel and Outlook
    excel = win32.Dispatch('Excel.Application')
    outlook = win32.Dispatch('Outlook.Application')
    excel.DisplayAlerts = False

    # loop through the specified number of times
    for i in range(num_emails):
        # open Excel and load the workbook
        ######## modifier : nom du fichier excel par la version récente ################################################################################
        workbook = excel.Workbooks.Open(r"C:\Users\Melek\Desktop\QS1.xlsx")

        # select the worksheet and range to copy
        ############# Modifier:  Nom worksheet par la dernière version et Range par le range des tableaux à envoyer ###########################################"
        worksheet = workbook.Worksheets('Sheet1')
        range_to_copy = worksheet.Range('K2:R17')

        # copy the range to the clipboard
        range_to_copy.Copy()

        # create a new Outlook email message
        mail = outlook.CreateItem(0)

        # Modifier les addresses mail et le sujet du mail
        mail.To = 'melekkatri8@gmail.com ; s.boutheina@ningen-groupe.com ; benhamouda.1@ningen-groupe.com ; o.bensalem@ningen-groupe.com ; mhirsi.1@ningen-groupe.com ; mejri.3@ningen-groupe.com ; Manel Madouri m.dph@ningen-groupe.com ; i.dridi@ningen-groupe.com ; m.hannafi@ningen-groupe.com ; m.bech@ningen-groupe.com ; r.benabdallah@ningen-groupe.com ; s.mrabet@ningen-groupe.com ; s.ducreux@ningen-groupe.com ; hassairi.1@ningen-groupe.com ; a.hamouda@ningen-groupe.com ; a.dridi@ningen-groupe.com'
        mail.Subject = 'Intraday ODS DZ'

        # paste the copied range into the email body
        mail.Display()
        inspector = mail.GetInspector
        editor = inspector.WordEditor
        editor.Windows(1).Selection.Paste()

        # send the email
        mail.Send()
        workbook.Close()
    return somme_ODS_DZ_recu,somme_ODS_DZ_tr

def PG_DZ(somme_PG_DZ_TR,somme_PG_DZ_recu,column_PG_DZ,column_PG_DZ_TR,row_PG_DZ):
    # Credentials to log in to the website
    username = 'sabrine.hassairi.adm'
    password = 'Ningen0925*'
    # URL of the website to scrape
    url = 'https://10.6.4.120/dashboards/custom/22'
    url1 = 'https://10.6.4.120/dashboards/custom/23'
    # Set up the Chrome driver service
    service = webdriver.chrome.service.Service('/Applications/driver/chromedriver')
    # Set up the webdriver
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--ignore-ssl-errors')
    driver = webdriver.Chrome(service=service, options=options)

    # Navigate to the login page
    driver.get('https://10.6.4.120/login')
    time.sleep(10)
    # insert the login details
    username_field = driver.find_element(By.XPATH,
                                         "/html/body/div[1]/div/md-content/div/div[1]/div/form/md-input-container[1]/input")
    username_field.send_keys(username)
    password_field = driver.find_element(By.XPATH,
                                         "/html/body/div[1]/div/md-content/div/div[1]/div/form/md-input-container[2]/input")
    password_field.send_keys(password)
    # Submit the login form
    password_field.send_keys(Keys.RETURN)
    # Wait for the login to complete
    time.sleep(10)

    # variables


    # Navigate to the URL
    driver.get(url)
    # Wait for the page to load
    time.sleep(15)

    xpaths = [
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[1]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[2]/md-card/md-card-content/div/div/div/div/div[1]"
    
    ]
    value1, value2 = [element.text for element in
                      driver.find_elements(By.XPATH, " | ".join(xpaths))]

    CS_PG_DZ = value1
    CS_PG_TR_DZ = value2

    print(CS_PG_DZ)
    print(CS_PG_TR_DZ)

    # Load the Excel file
    ######## modifier : nom du fichier excel par la version récente ######################################################################
    wf = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")

    # Select the worksheet you want to work with
    ws = wf.active
    # Select the cell to write in
    cell = ws.cell(row=row_PG_DZ, column=column_PG_DZ)
    lastval6 = int(CS_PG_DZ) - somme_PG_DZ_recu
    # write the new value
    cell.value = lastval6
    # new value of sum
    somme_PG_DZ_recu += lastval6

    # Select the cell to write in
    cell = ws.cell(row=row_PG_DZ, column=column_PG_DZ_TR)
    lastval6 = int(CS_PG_TR_DZ) - somme_PG_DZ_TR
    # write the new value
    cell.value = lastval6
    # new value of sum
    somme_PG_DZ_TR += lastval6


    ######## modifier : nom du fichier excel par la version récente#########################################################################
    wf.save(r"C:\Users\Melek\Desktop\QS1.xlsx")


    # set the number of times to send the email
    num_emails = 1

    # create Dispatch objects for Excel and Outlook
    excel = win32.Dispatch('Excel.Application')
    outlook = win32.Dispatch('Outlook.Application')
    excel.DisplayAlerts = False

    # loop through the specified number of times
    for i in range(num_emails):
        # open Excel and load the workbook
        ######## modifier : nom du fichier excel par la version récente ################################################################################
        workbook = excel.Workbooks.Open(r"C:\Users\Melek\Desktop\QS1.xlsx")

        # select the worksheet and range to copy
        ############# Modifier:  Nom worksheet par la dernière version et Range par le range des tableaux à envoyer ###########################################"
        worksheet = workbook.Worksheets('Sheet1')
        range_to_copy = worksheet.Range('A2:H14')

        # copy the range to the clipboard
        range_to_copy.Copy()

        # create a new Outlook email message
        mail = outlook.CreateItem(0)

        # set the recipients, subject, and body of the email
        mail.To = 'melekkatri8@gmail.com  ; s.boutheina@ningen-groupe.com ; benhamouda.1@ningen-groupe.com ; o.bensalem@ningen-groupe.com ; mhirsi.1@ningen-groupe.com ; mejri.3@ningen-groupe.com ; Manel Madouri m.dph@ningen-groupe.com ; i.dridi@ningen-groupe.com ; m.hannafi@ningen-groupe.com ; m.bech@ningen-groupe.com ; r.benabdallah@ningen-groupe.com ; s.mrabet@ningen-groupe.com ; s.ducreux@ningen-groupe.com ; hassairi.1@ningen-groupe.com ; a.hamouda@ningen-groupe.com ; a.dridi@ningen-groupe.com'
        mail.Subject = 'Intraday PG DZ'

        # paste the copied range into the email body
        mail.Display()
        inspector = mail.GetInspector
        editor = inspector.WordEditor
        editor.Windows(1).Selection.Paste()

        # send the email
        mail.Send()
        workbook.Close()
        return somme_PG_DZ_recu, somme_PG_DZ_TR

def ODS_TN(SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU,column_ODS_TN,column_ODS_TN_tr,row_ODS_TN):
    # Credentials to log in to the website
    username = 'nader.bejaoui'
    password = 'Paris;23?'
    # URL of the website to scrape
    url = 'https://10.10.3.111/dashboards/custom/11'
    # Set up the Chrome driver service
    service = webdriver.chrome.service.Service('/Applications/driver/chromedriver')
    # Set up the webdriver
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--ignore-ssl-errors')
    driver = webdriver.Chrome(service=service, options=options)
    # Navigate to the login page
    driver.get('https://10.10.3.111/')
    time.sleep(10)
    # insert the login details
    username_field = driver.find_element(By.XPATH,
                                         "/html/body/div[1]/div/md-content/div/div[1]/div/form/md-input-container[1]/input")
    username_field.send_keys(username)
    password_field = driver.find_element(By.XPATH,
                                         "/html/body/div[1]/div/md-content/div/div[1]/div/form/md-input-container[2]/input")
    password_field.send_keys(password)
    # Submit the login form
    password_field.send_keys(Keys.RETURN)
    # Wait for the login to complete
    time.sleep(10)
    # Navigate to the URL
    driver.get(url)
    # Wait for the page to load
    time.sleep(10)


    xpaths = [
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[1]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[6]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[11]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[2]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[9]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[14]/md-card/md-card-content/div/div/div/div/div[1]"
    ]
    value1, value2, value3, value4, value5, value6 = [element.text for element in
                                                      driver.find_elements(By.XPATH, " | ".join(xpaths))]

    CS_SALES_ODS = value5
    CS_SALES_ODS_TR = value6

    print(CS_SALES_ODS)
    print(CS_SALES_ODS_TR)

    ######## modifier : nom du fichier excel par la version récente####################################################################################
    wc = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")

    # Select the worksheet you want to work with
    ws = wc.active

    # Select the cell to write in
    cell = ws.cell(row=row_ODS_TN, column=column_ODS_TN)
    lastval4 = int(CS_SALES_ODS) - SOMME_ODS_TN_RECU
    # write the new value
    cell.value = lastval4
    # new value of sum
    SOMME_ODS_TN_RECU += lastval4

    # Select the cell to write in
    cell = ws.cell(row=row_ODS_TN, column=column_ODS_TN_tr)
    lastval4 = int(CS_SALES_ODS_TR) - SOMME_ODS_TN_TR
    # write the new value
    cell.value = lastval4
    # new value of sum
    SOMME_ODS_TN_TR += lastval4

    ######## modifier : nom du fichier excel par la version récente########################################################################################"
    wc.save(r"C:\Users\Melek\Desktop\QS1.xlsx")


    # set the number of times to send the email
    num_emails = 1

    # create Dispatch objects for Excel and Outlook
    excel = win32.Dispatch('Excel.Application')
    outlook = win32.Dispatch('Outlook.Application')
    excel.DisplayAlerts = False

    # loop through the specified number of times
    for i in range(num_emails):
        # open Excel and load the workbook

        ######## modifier : nom du fichier excel par la version récente #################################################################################
        workbook = excel.Workbooks.Open(r"C:\Users\Melek\Desktop\QS1.xlsx")
        ############# Modifier:  Nom worksheet par la dernière version et Range par le range des tableaux à envoyer########################################
        worksheet = workbook.Worksheets('Sheet1')
        range_to_copy = worksheet.Range('AQ2:AX16')

        # copy the range to the clipboard
        range_to_copy.Copy()

        # create a new Outlook email message
        mail = outlook.CreateItem(0)

        # set the recipients, subject, and body of the email
        mail.To = 'melekkatri8@gmail.com ; s.boutheina@ningen-groupe.com ; benhamouda.1@ningen-groupe.com ; o.bensalem@ningen-groupe.com ; mhirsi.1@ningen-groupe.com ; mejri.3@ningen-groupe.com ; Manel Madouri m.dph@ningen-groupe.com ; i.dridi@ningen-groupe.com ; m.hannafi@ningen-groupe.com ; m.bech@ningen-groupe.com ; r.benabdallah@ningen-groupe.com ; s.mrabet@ningen-groupe.com ; s.ducreux@ningen-groupe.com ; hassairi.1@ningen-groupe.com ; a.hamouda@ningen-groupe.com ; a.dridi@ningen-groupe.com'
        mail.Subject = 'Intraday ODS TN'

        # paste the copied range into the email body
        mail.Display()
        inspector = mail.GetInspector
        editor = inspector.WordEditor
        editor.Windows(1).Selection.Paste()

        # send the email
        mail.Send()
        workbook.Close()

    return SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU

def PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN):
    # PART2: WEB SCRAPPING
    # Credentials to log in to the website
    username = 'nader.bejaoui'
    password = 'Paris;23?'
    # URL of the website to scrape
    url = 'https://10.10.3.111/dashboards/custom/11'
    # Set up the Chrome driver service
    service = webdriver.chrome.service.Service('/Applications/driver/chromedriver')
    # Set up the webdriver
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--ignore-ssl-errors')
    driver = webdriver.Chrome(service=service, options=options)
    # Navigate to the login page
    driver.get('https://10.10.3.111/')
    time.sleep(10)
    # insert the login details
    username_field = driver.find_element(By.XPATH,
                                         "/html/body/div[1]/div/md-content/div/div[1]/div/form/md-input-container[1]/input")
    username_field.send_keys(username)
    password_field = driver.find_element(By.XPATH,
                                         "/html/body/div[1]/div/md-content/div/div[1]/div/form/md-input-container[2]/input")
    password_field.send_keys(password)
    # Submit the login form
    password_field.send_keys(Keys.RETURN)
    # Wait for the login to complete
    time.sleep(10)
    # Navigate to the URL
    driver.get(url)
    # Wait for the page to load
    time.sleep(10)
    xpaths = [
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[1]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[6]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[11]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[2]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[9]/md-card/md-card-content/div/div/div/div/div[1]",
        "/html/body/div[1]/div[2]/div/md-content/div[1]/md-content/ms-widget-engine/div/ul/li[14]/md-card/md-card-content/div/div/div/div/div[1]"
    ]
    value1, value2, value3, value4, value5, value6 = [element.text for element in
                                                      driver.find_elements(By.XPATH, " | ".join(xpaths))]

    CS_PG = value1
    CS_PG_TR = value2
    CS_SALES_PG = value3
    CS_SALES_PG_TR = value4

    print(CS_PG)


    ################################# modifier : nom du fichier excel par la version récente #########################################################################
    # Load the Excel file
    wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")

    # Select the worksheet you want to work with
    ws = wb.active
    

    # Select the cell to write in
    cell = ws.cell(row=row_PG_TN, column=column)
    lastval = int(CS_PG) - somme_PG_recu
    # write the new value
    cell.value = lastval
    # new value of sum
    somme_PG_recu += lastval
    # Select the cell to write in
    cell = ws.cell(row=row_PG_TN, column=column1)
    lastval1 = int(CS_PG_TR) - somme_PG_tr
    # write the new value
    cell.value = lastval1
    # new value of sum
    somme_PG_tr += lastval1
    # Select the cell to write in
    cell = ws.cell(row=row_PG_TN, column=column2)
    lastval2 = int(CS_SALES_PG) - somme_sales_recu
    # write the new value
    cell.value = lastval2
    # new value of sum
    somme_sales_recu += lastval2
    # Select the cell to write in
    cell = ws.cell(row=row_PG_TN, column=column3)
    lastval3 = int(CS_SALES_PG_TR) - somme_sales_tr
    # write the new value
    cell.value = lastval3
    # new value of sum
    somme_sales_tr += lastval3


    ################### modifier : nom du fichier excel par la version récente ###################################################################################
    wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")

    wb.close()

    # set the number of times to send the email
    num_emails = 1
    # create Dispatch objects for Excel and Outlook
    excel = win32.Dispatch('Excel.Application')
    outlook = win32.Dispatch('Outlook.Application')
    excel.DisplayAlerts = False
    # loop through the specified number of times
    for i in range(num_emails):
        # open Excel and load the workbook

        ######## modifier : nom du fichier excel par la version récente ##############################################################################
        workbook = excel.Workbooks.Open(r"C:\Users\Melek\Desktop\QS1.xlsx")

        ############# Modifier:  Nom worksheet par la dernière version et Range par le range des tableaux à envoyer####################################
        worksheet = workbook.Worksheets('Sheet1')
        range_to_copy = worksheet.Range('V2:AM16')

        # copy the range to the clipboard
        range_to_copy.Copy()
        # create a new Outlook email message
        mail = outlook.CreateItem(0)
        # set the recipients, subject, and body of the email
        mail.To = 'melekkatri8@gmail.com ; s.boutheina@ningen-groupe.com ; benhamouda.1@ningen-groupe.com ; o.bensalem@ningen-groupe.com ; mhirsi.1@ningen-groupe.com ; mejri.3@ningen-groupe.com ; Manel Madouri m.dph@ningen-groupe.com ; i.dridi@ningen-groupe.com ; m.hannafi@ningen-groupe.com ; m.bech@ningen-groupe.com ; r.benabdallah@ningen-groupe.com ; s.mrabet@ningen-groupe.com ; s.ducreux@ningen-groupe.com ; hassairi.1@ningen-groupe.com ; a.hamouda@ningen-groupe.com ; a.dridi@ningen-groupe.com'
        mail.Subject = 'Intraday PG TN'
        # paste the copied range into the email body
        mail.Display()
        inspector = mail.GetInspector
        editor = inspector.WordEditor
        editor.Windows(1).Selection.Paste()
        # send the email
        mail.Send()
        workbook.Close()
        return somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu

def test_vpn():
    while True:
        try:
            username = 'sabrine.hassairi.adm'
            password = 'Ningen0925*'
            # URL of the website to scrape
            url = 'https://10.6.4.120/dashboards/custom/22'
            url1 = 'https://10.6.4.120/dashboards/custom/23'
            # Set up the Chrome driver service
            service = webdriver.chrome.service.Service('/Applications/driver/chromedriver')
            # Set up the webdriver
            options = webdriver.ChromeOptions()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            driver = webdriver.Chrome(service=service, options=options)

            # Navigate to the login page
            driver.get('https://10.6.4.120/login')
            time.sleep(10)
            # insert the login details
            username_field = driver.find_element(By.XPATH,
                                                 "/html/body/div[1]/div/md-content/div/div[1]/div/form/md-input-container[1]/input")
            username_field.send_keys(username)
            password_field = driver.find_element(By.XPATH,
                                                 "/html/body/div[1]/div/md-content/div/div[1]/div/form/md-input-container[2]/input")
            password_field.send_keys(password)
            # Submit the login form
            password_field.send_keys(Keys.RETURN)
            return 0
        except WebDriverException as e:
            print("Erreur:", e )
            print("Attente d'une minute avant de réessayer...")
            time.sleep(60)

###########MAIN##########
i=0
#variables ODS DZ
somme_ODS_DZ_tr = 0
somme_ODS_DZ_recu = 0
column_ODS_DZ = 14
column_ODS_DZ_tr = 15
###################
row_ODS_DZ = 4

#variable PG dz
somme_PG_DZ_recu = 0
somme_PG_DZ_TR = 0
column_PG_DZ = 4
column_PG_DZ_TR = 5
#######################
row_PG_DZ = 4
somme_prevu_Inbound_Dz = 0
somme_prevu_ODS_Dz = 0

#variables ODS TN
SOMME_ODS_TN_TR = 0
SOMME_ODS_TN_RECU = 0
#### Ne pas Modifier
column_ODS_TN = 46
column_ODS_TN_tr = 47
#############################################
row_ODS_TN = 4


# variables PG TN
somme_PG_recu = 0
somme_PG_tr = 0
somme_sales_recu = 0
somme_sales_tr = 0

column = 25  
column1 = 26
column2 = 35
column3 = 36
#######################################
row_PG_TN = 4
somme_prevu_Inbound_Tn = 0



while True:
    today = datetime.date.today()
    thishour = datetime.datetime.now().time()
    
    if today.weekday() in [0 , 1 , 2 , 3]: #Monday to Thursday
          
        if thishour.hour == 10 and thishour.minute == 0:     
             
             
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles TN
            cell_total_prevu_Inbound_Tn.value = cell_cureent_Inbound_Tn.value
            somme_prevu_Inbound_Tn += cell_cureent_Inbound_Tn.value
            cell_total_prevu_Inbound_Tn.value = somme_prevu_Inbound_Tn
            # Global intervalles DZ
            cell_total_prevu_Inbound_Dz.value = cell_cureent_Inbound_Dz.value
            somme_prevu_Inbound_Dz += cell_cureent_Inbound_Dz.value
            cell_total_prevu_Inbound_Dz.value = somme_prevu_Inbound_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
             
           
             
            
            test_vpn()
            somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu=PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN)
            row_PG_TN+=1
            somme_PG_DZ_recu,somme_PG_DZ_TR=PG_DZ(somme_PG_DZ_TR,somme_PG_DZ_recu,column_PG_DZ,column_PG_DZ_TR,row_PG_DZ)
            row_PG_DZ += 1
            
        elif thishour.hour == 11 and thishour.minute == 0:
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles TN
            cell_total_prevu_Inbound_Tn.value = cell_cureent_Inbound_Tn.value
            somme_prevu_Inbound_Tn += cell_cureent_Inbound_Tn.value
            cell_total_prevu_Inbound_Tn.value = somme_prevu_Inbound_Tn
            # Global intervalles DZ
            cell_total_prevu_Inbound_Dz.value = cell_cureent_Inbound_Dz.value
            somme_prevu_Inbound_Dz += cell_cureent_Inbound_Dz.value
            cell_total_prevu_Inbound_Dz.value = somme_prevu_Inbound_Dz

            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            test_vpn()
            somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu=PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN)
            row_PG_TN+=1
            somme_PG_DZ_recu,somme_PG_DZ_TR=PG_DZ(somme_PG_DZ_TR,somme_PG_DZ_recu,column_PG_DZ,column_PG_DZ_TR,row_PG_DZ)
            row_PG_DZ += 1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1
        
        elif thishour.hour in range (12,19) and thishour.minute==0:
            
            
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles TN
            cell_total_prevu_Inbound_Tn.value = cell_cureent_Inbound_Tn.value
            somme_prevu_Inbound_Tn += cell_cureent_Inbound_Tn.value
            cell_total_prevu_Inbound_Tn.value = somme_prevu_Inbound_Tn
            # Global intervalles DZ
            cell_total_prevu_Inbound_Dz.value = cell_cureent_Inbound_Dz.value
            somme_prevu_Inbound_Dz += cell_cureent_Inbound_Dz.value
            cell_total_prevu_Inbound_Dz.value = somme_prevu_Inbound_Dz

            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            
            
            test_vpn()
            somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu=PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN)
            row_PG_TN+=1
            SOMME_ODS_TN_TR, SOMME_ODS_TN_RECU=ODS_TN(SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU,column_ODS_TN,column_ODS_TN_tr,row_ODS_TN)
            row_ODS_TN+= 1
            somme_PG_DZ_recu,somme_PG_DZ_TR=PG_DZ(somme_PG_DZ_TR,somme_PG_DZ_recu,column_PG_DZ,column_PG_DZ_TR,row_PG_DZ)
            row_PG_DZ += 1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1
        
        elif thishour.hour in range (19,21) and thishour.minute==0:
            
            
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles TN
            cell_total_prevu_Inbound_Tn.value = cell_cureent_Inbound_Tn.value
            somme_prevu_Inbound_Tn += cell_cureent_Inbound_Tn.value
            cell_total_prevu_Inbound_Tn.value = somme_prevu_Inbound_Tn
            # Global intervalles DZ
            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            test_vpn()
            somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu=PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN)
            row_PG_TN+=1
            SOMME_ODS_TN_TR, SOMME_ODS_TN_RECU=ODS_TN(SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU,column_ODS_TN,column_ODS_TN_tr,row_ODS_TN)
            row_ODS_TN+= 1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1
       
        elif thishour.hour in range (21,23) and thishour.minute==0:
            
            
            
                        #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        

            # Global intervalles DZ
            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            
            
            test_vpn()
            SOMME_ODS_TN_TR, SOMME_ODS_TN_RECU=ODS_TN(SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU,column_ODS_TN,column_ODS_TN_tr,row_ODS_TN)
            row_ODS_TN+= 1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1

        elif thishour.hour>23:
            break
        
    elif today.weekday() == 4: #Friday 
        row_ODS_DZ = 9
        if thishour.hour == 10 and thishour.minute == 0:
            
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles TN
            cell_total_prevu_Inbound_Tn.value = cell_cureent_Inbound_Tn.value
            somme_prevu_Inbound_Tn += cell_cureent_Inbound_Tn.value
            cell_total_prevu_Inbound_Tn.value = somme_prevu_Inbound_Tn
            

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            test_vpn()
            somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu=PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN)
            row_PG_TN+=1
            
        elif thishour.hour == 11 and thishour.minute == 0:
            
            
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles TN
            cell_total_prevu_Inbound_Tn.value = cell_cureent_Inbound_Tn.value
            somme_prevu_Inbound_Tn += cell_cureent_Inbound_Tn.value
            cell_total_prevu_Inbound_Tn.value = somme_prevu_Inbound_Tn
           

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            
            test_vpn()
            somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu=PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN)
            row_PG_TN+=1
        
        elif thishour.hour in range (12,15) and thishour.minute==0:
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles TN
            cell_total_prevu_Inbound_Tn.value = cell_cureent_Inbound_Tn.value
            somme_prevu_Inbound_Tn += cell_cureent_Inbound_Tn.value
            cell_total_prevu_Inbound_Tn.value = somme_prevu_Inbound_Tn
            

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            test_vpn()
            somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu=PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN)
            row_PG_TN+=1
            SOMME_ODS_TN_TR, SOMME_ODS_TN_RECU=ODS_TN(SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU,column_ODS_TN,column_ODS_TN_tr,row_ODS_TN)
            row_ODS_TN+= 1
           
        
        elif thishour.hour in range (15,21) and thishour.minute==0:
            
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles TN
            cell_total_prevu_Inbound_Tn.value = cell_cureent_Inbound_Tn.value
            somme_prevu_Inbound_Tn += cell_cureent_Inbound_Tn.value
            cell_total_prevu_Inbound_Tn.value = somme_prevu_Inbound_Tn
            # Global intervalles DZ
            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            
            test_vpn()
            somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu=PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN)
            row_PG_TN+=1
            SOMME_ODS_TN_TR, SOMME_ODS_TN_RECU=ODS_TN(SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU,column_ODS_TN,column_ODS_TN_tr,row_ODS_TN)
            row_ODS_TN+= 1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1
            
            
        elif thishour.hour in range (21,23) and thishour.minute==0:
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles DZ
            
            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            test_vpn()
            SOMME_ODS_TN_TR, SOMME_ODS_TN_RECU=ODS_TN(SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU,column_ODS_TN,column_ODS_TN_tr,row_ODS_TN)
            row_ODS_TN+= 1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1
            
        elif thishour.hour>23:
            break
        
        
    elif today.weekday() == 5: #Saturday 
        if thishour.hour == 10 and thishour.minute == 0:
            
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles TN
            cell_total_prevu_Inbound_Tn.value = cell_cureent_Inbound_Tn.value
            somme_prevu_Inbound_Tn += cell_cureent_Inbound_Tn.value
            cell_total_prevu_Inbound_Tn.value = somme_prevu_Inbound_Tn
            # Global intervalles DZ
            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            test_vpn()
            somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu=PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN)
            row_PG_TN+=1
            
        elif thishour.hour == 11 and thishour.minute == 0:
            
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles TN
            cell_total_prevu_Inbound_Tn.value = cell_cureent_Inbound_Tn.value
            somme_prevu_Inbound_Tn += cell_cureent_Inbound_Tn.value
            cell_total_prevu_Inbound_Tn.value = somme_prevu_Inbound_Tn
            # Global intervalles DZ
            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            
            test_vpn()
            somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu=PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN)
            row_PG_TN+=1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1
        
        elif thishour.hour in range (12,21) and thishour.minute==0:
            
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles TN
            cell_total_prevu_Inbound_Tn.value = cell_cureent_Inbound_Tn.value
            somme_prevu_Inbound_Tn += cell_cureent_Inbound_Tn.value
            cell_total_prevu_Inbound_Tn.value = somme_prevu_Inbound_Tn
            # Global intervalles DZ
            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            
            test_vpn()
            somme_sales_tr,somme_sales_recu,somme_PG_tr,somme_PG_recu=PG_TN(somme_PG_recu,somme_PG_tr,somme_sales_recu,somme_sales_tr,column,column1,column2,column3,row_PG_TN)
            row_PG_TN+=1
            SOMME_ODS_TN_TR, SOMME_ODS_TN_RECU=ODS_TN(SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU,column_ODS_TN,column_ODS_TN_tr,row_ODS_TN)
            row_ODS_TN+= 1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1
        
        elif thishour.hour in range (21,23) and thishour.minute==0:
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles DZ
            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            test_vpn()
            SOMME_ODS_TN_TR, SOMME_ODS_TN_RECU=ODS_TN(SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU,column_ODS_TN,column_ODS_TN_tr,row_ODS_TN)
            row_ODS_TN+= 1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1
                    
        elif thishour.hour>23:
            break
      
    else: #Sunday
        if thishour.hour == 10 and thishour.minute == 0:
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            
            # Global intervalles DZ
            cell_total_prevu_Inbound_Dz.value = cell_cureent_Inbound_Dz.value
            somme_prevu_Inbound_Dz += cell_cureent_Inbound_Dz.value
            cell_total_prevu_Inbound_Dz.value = somme_prevu_Inbound_Dz


            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            test_vpn()
            somme_PG_DZ_recu,somme_PG_DZ_TR=PG_DZ(somme_PG_DZ_TR,somme_PG_DZ_recu,column_PG_DZ,column_PG_DZ_TR,row_PG_DZ)
            row_PG_DZ += 1
            
        elif thishour.hour == 11 and thishour.minute == 0:
            
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            
            # Global intervalles DZ
            cell_total_prevu_Inbound_Dz.value = cell_cureent_Inbound_Dz.value
            somme_prevu_Inbound_Dz += cell_cureent_Inbound_Dz.value
            cell_total_prevu_Inbound_Dz.value = somme_prevu_Inbound_Dz

            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            
            test_vpn()
            somme_PG_DZ_recu,somme_PG_DZ_TR=PG_DZ(somme_PG_DZ_TR,somme_PG_DZ_recu,column_PG_DZ,column_PG_DZ_TR,row_PG_DZ)
            row_PG_DZ += 1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1
            
        elif thishour.hour in range (12,19) and thishour.minute==0:
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            
            # Global intervalles DZ
            cell_total_prevu_Inbound_Dz.value = cell_cureent_Inbound_Dz.value
            somme_prevu_Inbound_Dz += cell_cureent_Inbound_Dz.value
            cell_total_prevu_Inbound_Dz.value = somme_prevu_Inbound_Dz

            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            
            test_vpn()
            SOMME_ODS_TN_TR, SOMME_ODS_TN_RECU=ODS_TN(SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU,column_ODS_TN,column_ODS_TN_tr,row_ODS_TN)
            row_ODS_TN+= 1
            somme_PG_DZ_recu,somme_PG_DZ_TR=PG_DZ(somme_PG_DZ_TR,somme_PG_DZ_recu,column_PG_DZ,column_PG_DZ_TR,row_PG_DZ)
            row_PG_DZ += 1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1
            
        elif thishour.hour in range (19,23) and thishour.minute==0:
            
            #Global intervalle
            wb = openpyxl.load_workbook(r"C:\Users\Melek\Desktop\QS1.xlsx")
            # Select the worksheet you want to work with
            ws = wb.active          
            # Variables global intervalle
            cell_total_prevu_Inbound_Tn = ws.cell(row=15, column=24)
            cell_cureent_Inbound_Tn = ws.cell(row=row_PG_TN, column=24)
            
            cell_total_prevu_Inbound_Dz = ws.cell(row=13, column=3)
            cell_cureent_Inbound_Dz = ws.cell(row=row_PG_DZ, column=3)
            
            cell_total_prevu_ODS_Dz = ws.cell(row=16, column=13)
            cell_cureent_ODS_Dz = ws.cell(row=row_ODS_DZ, column=13)        
            # Global intervalles DZ
            cell_total_prevu_ODS_Dz.value = cell_cureent_ODS_Dz.value
            somme_prevu_ODS_Dz += cell_cureent_ODS_Dz.value
            cell_total_prevu_ODS_Dz.value = somme_prevu_ODS_Dz

            wb.save(r"C:\Users\Melek\Desktop\QS1.xlsx")
            
            
            test_vpn()
            SOMME_ODS_TN_TR, SOMME_ODS_TN_RECU=ODS_TN(SOMME_ODS_TN_TR,SOMME_ODS_TN_RECU,column_ODS_TN,column_ODS_TN_tr,row_ODS_TN)
            row_ODS_TN+= 1
            somme_ODS_DZ_recu,somme_ODS_DZ_tr=ODS_DZ(somme_ODS_DZ_tr,somme_ODS_DZ_recu,column_ODS_DZ,column_ODS_DZ_tr,row_ODS_DZ)
            row_ODS_DZ += 1 
            
        elif thishour.hour>23:
            break
        
        
